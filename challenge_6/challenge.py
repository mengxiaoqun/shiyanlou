# -*- coding:utf-8 -*- 

from openpyxl import load_workbook
from openpyxl import Workbook
import datetime

wb = load_workbook('courses.xlsx')
student_sheet = wb.get_sheet_by_name('students')
time_sheet = wb.get_sheet_by_name('time')


def combine():

	combine_sheet = wb.create_sheet()
	combine_sheet.title = 'combine'
	combine_sheet.append(['创建时间','课程名称','学习人数','学习时间'])
	

	for stu in student_sheet.values:
		if stu[2] != '学习人数':
			for time in time_sheet.values:
				if time[1] == stu[1]:
					combine_sheet.append(list(stu)+[time[2]])
	wb.save('courses.xlsx')


def split():

	combine_sheet = wb.get_sheet_by_name('combine')
	time_list = []
	for combi in combine_sheet.values:
		if combi[0] != '创建时间':
			time_list.append(combi[0].strftime('%Y'))

	for value in set(time_list):
		wb_tmp = Workbook()
		wb_tmp.remove(wb_tmp.active)
		ws = wb_tmp.create_sheet(title = value)
		ws.append(['创建时间','课程名称','学习人数','学习时间'])
	
		for combi in combine_sheet.values:
			if combi[0] != '创建时间':
				if value == combi[0].strftime('%Y'):
					ws.append(combi)

		wb_tmp.save('{}.xlsx'.format(value))



if __name__ == '__main__':
	combine()
	split()
